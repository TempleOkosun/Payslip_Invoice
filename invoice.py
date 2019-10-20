# Requirements
import openpyxl  # Library needed to read/write excel files

# Library needed for generating pdfs and graphics.
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from PIL import Image

from functions import merge_pdfs

# Setting to support the use of Arial font
pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))

# Load the excel document and sheet
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb['invoices']

# Set the page size A4
page_width = 2480
page_height = 3508
margin = 100


# Specify logo
im = Image.open('Pynet_Logo.jpg')
width, height = im.size
ratio = width/height
image_width = 400
image_height = image_width / ratio

# Invoice related information
company_name = 'PyNet Technologies Limited'
month_year = 'August 2019'


def create_invoices():
    # Get last non empty row in the sheet
    last_row = sheet.max_row
    for i in range(2, last_row + 1):
        customer = sheet.cell(row=i, column=3).value
        invoice_number = sheet.cell(row=i, column=1).value
        invoice_date = sheet.cell(row=i, column=8).value
        due_date = sheet.cell(row=i, column=9).value
        description = sheet.cell(row=i, column=4).value
        amount_excl_vat = sheet.cell(row=i, column=5).value
        vat = sheet.cell(row=i, column=6).value
        total_amount = sheet.cell(row=i, column=7).value

        # Set what the file is saved as
        c = canvas.Canvas(str(invoice_number) + "_" + customer + ".pdf")

        # Set the page size
        c.setPageSize((page_width, page_height))

        # Draw the logo
        c.drawInlineImage('Pynet_logo.jpg', page_width - image_width - margin, page_height - image_height, image_width,
                          image_width)

        # Format the heading text 'Invoice'.
        c.setFont('Arial', 80)
        text = 'INVOICE'
        text_width = stringWidth(text, 'Arial', 80)
        c.drawString((page_width - text_width) / 2, page_height - image_height - margin, text)

        # Formatting the margin for proper label and actual content positioning on the page.
        start = 2 * margin  # start- initialized before function def., tells margin before label.
        start_2 = start + 500  # start_2 initialized above too tells us margin to leave before content.
        y = page_height - image_height - margin * 2.5  # Sets the page level/height we will start contents from.

        # Set font size for text hence forth i.e. the page body.
        c.setFont('Arial', 45)

        # Company name
        c.drawString(start, y, 'Issued by: ')
        c.drawString(start_2, y, company_name)
        y -= margin

        # Customer
        c.drawString(start, y, 'Issued to: ')
        c.drawString(start_2, y, customer)
        y -= margin

        # Invoice number
        c.drawString(start, y, 'Invoice number: ')
        c.drawString(start_2, y, str(invoice_number))
        y -= margin

        # Invoice date
        c.drawString(start, y, 'Issued date: ')
        c.drawString(start_2, y, str(invoice_date))
        y -= margin

        # Due date
        c.drawString(start, y, 'Due date: ')
        c.drawString(start_2, y, due_date)
        y -= margin * 2

        # Description
        c.drawString(start, y, 'Invoice issued for performed: ' + description + ' for ' + month_year)
        y -= margin * 2

        # Amount excl. VAT
        c.drawString(start, y, 'Amount excluding VAT: ')
        c.drawString(start_2, y, 'CAD ' + str(amount_excl_vat))
        y -= margin

        # VAT
        c.drawString(start, y, 'Value Added Tax: ')
        c.drawString(start_2, y, 'CAD ' + str(vat))
        y -= margin

        # Total amount
        c.drawString(start, y, 'Total Amount: ')
        c.drawString(start_2, y, 'CAD ' + str(total_amount))
        y -= margin * 4

        # Payment terms
        c.drawString(start, y, 'If paid within 10 days, 2% of discount is granted.')
        y -= margin

        # Account details
        c.drawString(start, y, 'Bank account number: 1234 ABCD 5670 Missisauga, ON, Canada.')
        y -= margin

        # Contact
        c.drawString(start, y, 'For any information, please contact info@pynet.com')
        y -= margin

        c.save()


create_invoices()
merge_pdfs()