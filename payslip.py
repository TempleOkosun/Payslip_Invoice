# Requirements
import openpyxl  # Library needed to read/write excel files

# Library needed for generating pdfs and graphics.
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

from functions import merge_pdfs

# Setting to support the use of Arial font
pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))

# Load the excel document and sheet
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb['employees']

# Set the page size A4
page_width = 2480
page_height = 3508


month_year = "August 2019"
company_name = "PyNet Technologies Limited"
spread = 100
start = 200
start_2 = 700


def create_payslip():
    # Get last non empty row in the sheet
    last_row = sheet.max_row

    # Iterate over all the rows extracting the data.
    for i in range(2, last_row + 1):
        emp_id = sheet.cell(row=i, column=1).value
        emp_name = sheet.cell(row=i, column=2).value
        emp_last_name = sheet.cell(row=i, column=3).value
        gross_salary = sheet.cell(row=i, column=4).value
        pension_contribution = sheet.cell(row=i, column=5).value
        health_insurance = sheet.cell(row=i, column=6).value
        personal_income_tax = sheet.cell(row=i, column=7).value
        bonus_pay = sheet.cell(row=i, column=8).value
        deduction = sheet.cell(row=i, column=9).value
        net_salary = sheet.cell(row=i, column=10).value

        # Set what the file is saved as
        c = canvas.Canvas(str(emp_name) + '_' + str(emp_last_name) + '_' + str(emp_id) + '_' + month_year + '.pdf')

        # Set the page size
        c.setPageSize((page_width, page_height))

        # Format the heading texts Company name and period.
        c.setFont('Arial', 100)
        company_text_width = stringWidth( company_name, 'Arial', 100)
        c.drawString((page_width - company_text_width) / 2, 3300, company_name)

        c.setFont('Arial', 50)
        text = "Salary calculation for period " + month_year
        text_width = stringWidth(text, 'Arial', 50)
        c.drawString((page_width - text_width) / 2, 3120, text)

        # The actual page body content starts here.
        y = 3000  # Sets the page level/height we will start contents from.
        c.setFont('Arial', 45)
        c.drawString(start, y, "Employee's id: ")  # start- initialized before function def., tells margin before label.
        c.drawString(start_2, y, str(emp_id))  # start_2 initialized above too tells us margin to leave before content.
        y -= spread  # Line spacing for the next content. i.e sets where the next content will be on the page.

        c.drawString(start, y, "Employee's name: ")
        c.drawString(start_2, y, str(emp_name) + ' ' + str(emp_last_name))
        y -= spread

        c.drawString(start, y, "Gross salary: ")
        c.drawString(start_2, y, str(gross_salary))
        y -= spread

        c.drawString(start, y, "Pension contribution: ")
        c.drawString(start_2, y, str(pension_contribution))
        y -= spread

        c.drawString(start, y, "Health insurance: ")
        c.drawString(start_2, y, str(health_insurance))
        y -= spread

        c.drawString(start, y, "Personal income tax: ")
        c.drawString(start_2, y, str(personal_income_tax))
        y -= spread

        c.drawString(start, y, "Bonus pay: ")
        c.drawString(start_2, y, str(bonus_pay))
        y -= spread

        c.drawString(start, y, "Deduction: ")
        c.drawString(start_2, y, str(deduction))
        y -= spread

        c.drawString(start, y, "Net salary: ")
        c.drawString(start_2, y, str(net_salary))
        y -= spread  * 4

        c.drawString(start, y, 'Signature: ')
        c.drawString(start_2, y, '____________________')

        c.save()  # Saves the page


create_payslip()
merge_pdfs()
